#!/usr/bin/env python3
"""Render one real screenshot per workbook, plus copy in each Gumroad cover.

The cover.png that ships with each product is a branded mockup with illustrative
charts — good as an og:image, but it is not evidence of what is in the file. A
buyer deciding on a $119 workbook wants to see the actual thing. So this renders
the product's most persuasive tab straight from the shipped .xlsx.

How: copy the workbook, drop every sheet but the one we want, force fit-to-width
landscape, convert with headless LibreOffice, rasterise, trim the margins.

The copy is why this is safe. None of these workbooks contain charts or images
(verified: no xl/charts/* or xl/media/* in any of them), so an openpyxl
round-trip loses nothing that shows up in a picture — and it never touches the
shipped file either way.

Run:  python3 shots.py           # all 21
      python3 shots.py wip       # just the ones whose key contains "wip"
"""
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageChops

from catalog import PRODUCTS

SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
OUT = Path.home() / "Projects/allanninal.dev/spreadsheets/assets/img"
DPI = 150
MAX_W = 1600
SLACK_COLS = 3

# The tab to photograph: the one that shows the product doing its job, not the
# instructions tab. Falls back to the second sheet, since sheet 0 is always
# "Start Here".
HERO = {
    "construction-wip-schedule-workbook": "WIP Schedule",
    "construction-cash-flow-forecast": "Cash Flow",
    "progress-billing-schedule-of-values": "Application",
    "equipment-fleet-cost-per-hour": "Rate Build-Up",
    "certified-payroll-davis-bacon": "Annualization",
    "federal-grant-budget-mtdc": "Budget by Year",
    "landed-cost-duty-calculator": "Landed Cost",
    "spc-control-chart-capability-workbook": "Capability",
    "monte-carlo-simulation-workbook": "Results",
    "design-of-experiments-workbook": "Effects",
    "recipe-costing-menu-engineering": "Menu Engineering",
    "self-storage-underwriting-model": "Returns",
    "car-wash-underwriting-model": "Returns",
    "rv-park-underwriting-model": "Returns",
    "mobile-home-park-underwriting-model": "Returns",
    "laundromat-underwriting-model": "Returns",
    "electrical-bid-calculator": "Bid Summary",
    "hvac-bid-calculator": "Bid Summary",
    "roofing-bid-calculator": "Bid Summary",
    "concrete-bid-calculator": "Bid Summary",
    "landscaping-bid-calculator": "Bid Summary",
}


def isolate(src: Path, sheet: str, dst: Path) -> str:
    """Leave only `sheet` visible. HIDE the others, never delete them.

    Deleting was the first attempt and it renders a screenshot of #NAME? errors:
    every one of these workbooks pulls its inputs across tabs, so a lone sheet
    has nothing to reference. Hidden sheets still hold their values and still
    calculate, and LibreOffice omits them from the PDF.
    """
    wb = load_workbook(src)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"{src.name}: no sheet named {sheet!r} (have {wb.sheetnames})")
    for name in wb.sheetnames:
        wb[name].sheet_state = "visible" if name == sheet else "hidden"
    wb.active = wb.sheetnames.index(sheet)
    ws = wb[sheet]
    _widen(ws)
    last = _last_used_row(ws)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    # Long labels in the last used column overflow into the empty columns beside
    # it. Those columns are outside the used range, so without slack the print
    # area cuts the sentence in half.
    slack = ws.max_column + SLACK_COLS
    for col in range(ws.max_column + 1, slack + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.print_area = f"A1:{get_column_letter(slack)}{last}"
    wb.save(dst)
    return sheet


def _widen(ws) -> None:
    """Give every column an explicit width wide enough for what is in it.

    The summary cards on several of these sheets sit in columns the author never
    set a width for, so Excel's ~8-character default applies and LibreOffice
    prints a big currency figure as `###`. Harmless on screen where the user can
    drag a column; fatal in a screenshot.
    """
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].width:
            continue
        widest = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            # A formula's text is not what gets printed; assume a wide result.
            widest = max(widest, 16 if str(v).startswith("=") else len(str(v)))
        if widest:
            ws.column_dimensions[letter].width = min(max(widest + 3, 12), 32)


def _last_used_row(ws) -> int:
    """Last row holding anything, so blank filler rows do not pad the picture."""
    for row in range(ws.max_row, 0, -1):
        if any(ws.cell(row=row, column=c).value not in (None, "")
               for c in range(1, ws.max_column + 1)):
            return row
    return ws.max_row


def to_pdf(xlsx: Path, outdir: Path) -> Path:
    # A private profile dir keeps this from colliding with a running LibreOffice.
    subprocess.run(
        [SOFFICE, "--headless", "--norestore",
         f"-env:UserInstallation=file://{outdir}/profile",
         "--convert-to", "pdf", "--outdir", str(outdir), str(xlsx)],
        capture_output=True, text=True, timeout=180)
    pdf = outdir / (xlsx.stem + ".pdf")
    if not pdf.exists():
        raise SystemExit(f"LibreOffice produced no PDF for {xlsx.name}")
    return pdf


def trim(img: Image.Image) -> Image.Image:
    """Crop the page margins off, leaving a small even border."""
    rgb = img.convert("RGB")
    bg = Image.new("RGB", rgb.size, (255, 255, 255))
    box = ImageChops.difference(rgb, bg).getbbox()
    if not box:
        return img
    pad = 14
    l, t, r, b = box
    return img.crop((max(0, l - pad), max(0, t - pad),
                     min(img.width, r + pad), min(img.height, b + pad)))


def shoot(p: dict) -> None:
    key = p["key"]
    dest = OUT / key
    dest.mkdir(parents=True, exist_ok=True)

    shutil.copy2(p["cover"], dest / "cover.png")

    src = Path(p["xlsx"])
    sheet = HERO.get(key) or p["tabs"][1]
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        one = td / f"{key}.xlsx"
        isolate(src, sheet, one)
        pdf = to_pdf(one, td)
        subprocess.run(["pdftoppm", "-png", "-r", str(DPI), "-f", "1", "-l", "1",
                        str(pdf), str(td / "page")], check=True, capture_output=True)
        pages = sorted(td.glob("page*.png"))
        if not pages:
            raise SystemExit(f"{key}: pdftoppm produced nothing")
        img = trim(Image.open(pages[0]))
        if img.width > MAX_W:
            img = img.resize((MAX_W, round(img.height * MAX_W / img.width)),
                             Image.LANCZOS)
        img.convert("RGB").save(dest / "tab.png", optimize=True)

    kb = (dest / "tab.png").stat().st_size / 1024
    print(f"  ok  {key:<40} {sheet:<18} {img.width}x{img.height}  {kb:,.0f} KB")


def main() -> None:
    want = sys.argv[1] if len(sys.argv) > 1 else ""
    todo = [p for p in PRODUCTS if want in p["key"]]
    print(f"rendering {len(todo)} workbook screenshots -> {OUT}")
    for p in todo:
        shoot(p)


if __name__ == "__main__":
    main()
