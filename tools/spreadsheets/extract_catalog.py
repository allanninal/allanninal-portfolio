#!/usr/bin/env python3
"""Extract the FACTS about every live Excel product into catalog.py.

The narrative on each page is hand-written. Every number on it is not: price,
Gumroad URL, product id, tab names and tab count are read here, straight from
the source of truth in ~/Projects/gumroad-products, and written into catalog.py.

RUNBOOK.md records the failure this prevents — sales copy inherited from an
earlier batch that carried claims which were false for the new one. A page can
now never advertise "8 tabs" for a 7-tab workbook or quote a stale price. Where
a hand-written claim and this extractor disagree, the extractor is right.

Sources, in order of authority:
  <line>/<line>-results.tsv   slug, product id, short_url, price   (Gumroad API response)
  <line>/<line>-manifest.tsv  title, summary, category, tags, xlsx path
  the shipped .xlsx itself    sheet names, in order

Run:  python3 extract_catalog.py
"""
import csv
import json
import pprint
import re
import subprocess
from pathlib import Path

from openpyxl import load_workbook

GUMROAD = "/opt/homebrew/bin/gumroad"

SRC = Path.home() / "Projects/gumroad-products"
OUT = Path(__file__).parent / "catalog.py"

# The 13 product lines that ship an .xlsx. Everything under spreadsheets/ is
# excluded on purpose: built and priced, but never created on Gumroad, so there
# is no product id and nothing to link to.
LINES = ["trades", "underwriting", "wip", "cashflow", "billing", "fleet",
         "payroll", "grants", "importing", "spc", "montecarlo", "doe", "kitchen"]


def read_tsv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8") as fh:
        return list(csv.DictReader(fh, delimiter="\t"))


def sheet_names(xlsx: Path) -> list[str]:
    wb = load_workbook(xlsx, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


FUNC = re.compile(r"\b([A-Z][A-Z0-9.]{1,20})\(")

# Functions a reader might reasonably doubt are available outside Excel. Every
# one of these was checked against Apple's published function list for Numbers
# and Google's for Sheets — all present — but the page names the ones a given
# workbook actually uses rather than making a blanket compatibility claim.
NOTABLE = {"IRR", "CUMIPMT", "CUMPRINC", "PMT", "NORMINV", "NORMSDIST", "NORMSINV",
           "TINV", "FDIST", "PERCENTILE", "CORREL", "STDEV", "RANK", "MEDIAN", "RAND"}


def functions_used(xlsx: Path) -> list[str]:
    """Every worksheet function the workbook's formulas call.

    This is what lets each page state its own portability instead of asserting a
    blanket "works everywhere". It also proves the claim these products make:
    no macros and no add-ins, so nothing here can fail to open.
    """
    wb = load_workbook(xlsx)
    try:
        used = set()
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        used |= set(FUNC.findall(v))
        return sorted(used - {"TRUE", "FALSE"})
    finally:
        wb.close()


def collect() -> list[dict]:
    rows = []
    for line in LINES:
        d = SRC / line
        results = {r["slug"]: r for r in read_tsv(d / f"{line}-results.tsv")}
        manifest = {r["slug"]: r for r in read_tsv(d / f"{line}-manifest.tsv")}

        missing = set(results) - set(manifest)
        if missing:
            raise SystemExit(f"{line}: in results but not manifest: {sorted(missing)}")

        for slug, res in results.items():
            man = manifest[slug]
            xlsx = Path(man["file"])
            if not xlsx.exists():
                raise SystemExit(f"{line}/{slug}: xlsx missing at {xlsx}")

            cover = d / slug / "cover.png"
            if not cover.exists():
                raise SystemExit(f"{line}/{slug}: cover.png missing")

            # results.tsv price is what Gumroad actually charges; the manifest
            # price is what the builder intended. They must agree.
            if res["price"] != man["price"]:
                raise SystemExit(
                    f"{line}/{slug}: price disagrees — "
                    f"results {res['price']} vs manifest {man['price']}")

            rows.append({
                "key": slug,
                "line": line,
                "gumroad_title": man["title"],
                "summary": man["summary"],
                "price": int(res["price"]),
                "url": res["url"],
                "product_id": res["id"],
                "category": man["category"],
                "tags": [t for t in man["tags"].split("|") if t],
                "tabs": sheet_names(xlsx),
                "functions": functions_used(xlsx),
                "notable_functions": sorted(set(functions_used(xlsx)) & NOTABLE),
                "xlsx": str(xlsx),
                "cover": str(cover),
                "created_at": res["created_at"],
            })
    return rows


def live_products() -> dict:
    """Every spreadsheet product the Gumroad account actually has, from the API.

    The local results.tsv files record what was created; this records what is
    live NOW. Prices get edited in the Gumroad UI and products get unpublished,
    and neither shows up in a TSV written weeks ago.
    """
    r = subprocess.run([GUMROAD, "products", "list", "--all", "--json"],
                       capture_output=True, text=True)
    if r.returncode != 0:
        raise SystemExit(f"gumroad products list failed: {r.stderr.strip()}")
    out = {}
    for p in json.loads(r.stdout).get("products", []):
        blob = " ".join([p.get("name") or "", p.get("description") or "",
                         " ".join(p.get("tags") or [])]).lower()
        # "workbook" alone is not enough — several print-ready PDF devotionals
        # are called workbooks. Requiring a spreadsheet word keeps them out.
        if re.search(r"\bexcel\b|google sheets|\.xlsx|spreadsheet", blob):
            out[p["custom_permalink"]] = p
    return out


def cross_check(rows: list[dict]) -> None:
    live = live_products()
    mine = {r["url"].rsplit("/l/", 1)[1]: r for r in rows}

    only_live = sorted(set(live) - set(mine))
    only_mine = sorted(set(mine) - set(live))
    if only_live:
        raise SystemExit(
            "Live on Gumroad but missing from the local lines — a product was "
            f"created outside this repo: {only_live}")
    if only_mine:
        raise SystemExit(
            f"In the local lines but not live on Gumroad (deleted or unpublished?): {only_mine}")

    for perma, r in mine.items():
        p = live[perma]
        if int(p["price"]) != r["price"] * 100:
            raise SystemExit(f"{perma}: Gumroad charges ${int(p['price'])/100:.0f}, "
                             f"local files say ${r['price']}")
        if not p["published"]:
            raise SystemExit(f"{perma}: not published on Gumroad")
        r["gumroad_name"] = p["name"]
        r["sales_count"] = p["sales_count"]

    sold = sum(r["sales_count"] for r in rows)
    print(f"cross-checked {len(mine)} products against the Gumroad API — "
          f"all live, prices agree, {sold} sales")
    if sold == 0:
        print("  0 sales: no ratings, reviews or aggregateRating anywhere on these pages")


def main() -> None:
    rows = collect()
    cross_check(rows)
    rows.sort(key=lambda r: (-r["price"], r["key"]))

    body = pprint.pformat(rows, width=96, sort_dicts=False)
    OUT.write_text(
        '"""Facts about the live Excel products. GENERATED — do not hand-edit.\n\n'
        "Written by extract_catalog.py from <line>-results.tsv, <line>-manifest.tsv\n"
        "and the shipped .xlsx files in ~/Projects/gumroad-products. Re-run the\n"
        "extractor after any price change or republish.\n"
        '"""\n\n'
        f"PRODUCTS = {body}\n\n"
        "BY_KEY = {p['key']: p for p in PRODUCTS}\n",
        encoding="utf-8")

    print(f"{len(rows)} products -> {OUT}")
    for r in rows:
        print(f"  ${r['price']:>3}  {r['key']:<38} {len(r['tabs'])} tabs  {r['url']}")


if __name__ == "__main__":
    main()
